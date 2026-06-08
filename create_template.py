import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Sheet 1: Data Entry Form ──────────────────────────────────────────────────
ws = wb.active
ws.title = "Use Cases"

HEADER_FILL  = PatternFill("solid", fgColor="1E293B")
REQUIRED_FILL = PatternFill("solid", fgColor="FEF3C7")
OPTIONAL_FILL = PatternFill("solid", fgColor="F0FDF4")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
NOTE_FONT    = Font(italic=True, color="64748B", size=9)
BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

# Column definitions: (header, width, required, note)
columns = [
    ("ID",                        8,  False, "Leave blank — auto-assigned on import"),
    ("Department",               22,  True,  "Pick from dropdown"),
    ("Use Case Title",           35,  True,  "Short name, e.g. 'Predictive Maintenance'"),
    ("Goal",                     35,  True,  "One sentence: what business problem does it solve?"),
    ("KPIs",                     28,  True,  "Comma-separated metrics, e.g. 'Downtime%, OEE'"),
    ("Results",                  35,  True,  "Quantified outcome, e.g. '25-50% downtime↓'"),
    ("Example",                  60,  False, "Real company examples with year & numbers"),
    ("Fit Criteria",             50,  True,  "When is this use case a good fit?"),
    ("Not Fit When",             40,  True,  "When should you NOT use this?"),
    ("Time to Value",            18,  True,  "e.g. '6-9 months'"),
    ("Time to Value Min (mo)",   22,  True,  "Numeric minimum months"),
    ("Time to Value Max (mo)",   22,  True,  "Numeric maximum months"),
    ("Cost Tier",                14,  True,  "$ / $$ / $$$"),
    ("Strategic Goals",          45,  True,  "Comma-separated from allowed list"),
    ("Is AI4M?",                 14,  False, "TRUE / FALSE"),
    ("AI4M Companies",           35,  False, "Comma-separated company names"),
    ("Canadian Context?",        20,  False, "TRUE / FALSE"),
    ("Sustainability Focus?",    22,  False, "TRUE / FALSE"),
    ("Source / Reference URL",   50,  False, "URL or citation for the data"),
    ("Notes",                    50,  False, "Any extra context for reviewers"),
]

# Header row
for col_idx, (header, width, required, note) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Note row (row 2)
for col_idx, (header, width, required, note) in enumerate(columns, start=1):
    cell = ws.cell(row=2, column=col_idx, value=note)
    cell.fill = PatternFill("solid", fgColor="F1F5F9")
    cell.font = NOTE_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = BORDER

ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 40

# Pre-fill 50 data rows with alternating fills
for row in range(3, 53):
    for col_idx, (header, width, required, note) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 55

# ── Dropdowns ─────────────────────────────────────────────────────────────────
departments = [
    "Customer Service / After-Sales",
    "Engineering & R&D",
    "Finance & Administration",
    "Human Resources",
    "Information Technology",
    "Marketing",
    "Production & Operations",
    "Quality & Maintenance",
    "Sales",
    "Supply Chain & Procurement",
]
dept_formula = '"' + ','.join(departments) + '"'
dv_dept = DataValidation(type="list", formula1=dept_formula, allow_blank=True)
dv_dept.sqref = f"B3:B52"
ws.add_data_validation(dv_dept)

cost_dv = DataValidation(type="list", formula1='"$,$$,$$$"', allow_blank=True)
cost_dv.sqref = "M3:M52"
ws.add_data_validation(cost_dv)

bool_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
bool_dv.sqref = "O3:O52"
ws.add_data_validation(bool_dv)

bool_dv2 = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
bool_dv2.sqref = "Q3:Q52"
ws.add_data_validation(bool_dv2)

bool_dv3 = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
bool_dv3.sqref = "R3:R52"
ws.add_data_validation(bool_dv3)

# Freeze header rows
ws.freeze_panes = "A3"

# ── Sheet 2: Reference ────────────────────────────────────────────────────────
ref = wb.create_sheet("Reference Lists")
ref.column_dimensions["A"].width = 30
ref.column_dimensions["B"].width = 40

ref_headers = ["Strategic Goals (allowed values)", "Departments"]
for col, h in enumerate(ref_headers, start=1):
    c = ref.cell(row=1, column=col, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center")

strategic_goals = [
    "Cost Reduction", "Customer Experience", "ESG & Sustainability",
    "Innovation & Agility", "Operational Efficiency", "Quality & Compliance",
    "Revenue & Sales Growth", "Security & Risk", "Supply Chain & Inventory",
    "Workforce & Safety",
]
for i, g in enumerate(strategic_goals, start=2):
    ref.cell(row=i, column=1, value=g)

for i, d in enumerate(departments, start=2):
    ref.cell(row=i, column=2, value=d)

# ── Sheet 3: Example Row ──────────────────────────────────────────────────────
ex = wb.create_sheet("Example (Do Not Edit)")
ex_headers = [c[0] for c in columns]
for col_idx, h in enumerate(ex_headers, start=1):
    c = ex.cell(row=1, column=col_idx, value=h)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    ex.column_dimensions[get_column_letter(col_idx)].width = columns[col_idx-1][1]

example_row = [
    "",                                          # ID
    "Production & Operations",                   # Department
    "Predictive Maintenance",                    # Use Case Title
    "Reduce unplanned equipment downtime",       # Goal
    "Downtime%, OEE, MTBR",                     # KPIs
    "25-50% downtime reduction",                # Results
    "BMW Regensburg 2023: saved 500min/yr",      # Example
    "Vibration & temperature sensors; CMMS data; IoT history",  # Fit Criteria
    "Sparse sensor history; reactive-only culture",             # Not Fit When
    "6-9 months",                               # Time to Value
    6,                                          # Min months
    9,                                          # Max months
    "$$",                                       # Cost Tier
    "Operational Efficiency, Cost Reduction",   # Strategic Goals
    "TRUE",                                     # Is AI4M
    "Maya HTT, Canvass Analytics",              # AI4M Companies
    "FALSE",                                    # Canadian Context
    "FALSE",                                    # Sustainability Focus
    "https://example.com/case-study",           # Source URL
    "",                                         # Notes
]
for col_idx, val in enumerate(example_row, start=1):
    ex.cell(row=2, column=col_idx, value=val)

out_path = "/Users/fahadhafeez/Desktop/ai-usecase-explorer/AI_UseCase_Collection_Template.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
